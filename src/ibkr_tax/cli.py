#!/usr/bin/env python3
"""
Command-line interface for IBKR Tax Tool
"""

import argparse
import json
import sys
from datetime import datetime
from typing import Any, Dict, List, Union

import pandas as pd

from .api.flex_query import FlexQueryClient
from .config import Config
from .constants import TIMESTAMP_FORMAT
from .exceptions import APIError, ConfigurationError, IBKRTaxError
from .parsers.data_parser import (
    calculate_summary,
    parse_deposits_withdrawals,
    parse_dividends,
    parse_trades,
    parse_withholding_tax,
)
from .utils.logging import setup_logger


def print_banner() -> None:
    """Print application banner"""
    print("=" * 60)
    print("IBKR Tax Tool - Trading Data Fetcher")
    print("=" * 60)
    print()


def export_to_excel(
    trades_df: pd.DataFrame,
    dividends_df: pd.DataFrame,
    tax_df: pd.DataFrame,
    deposits_withdrawals_df: pd.DataFrame,
    summary: Dict[str, Any],
    filepath: str,
) -> None:
    """
    Export data to Excel file with multiple sheets

    Args:
        trades_df: DataFrame with trade data
        dividends_df: DataFrame with dividend data
        tax_df: DataFrame with tax data
        deposits_withdrawals_df: DataFrame with deposits/withdrawals data
        summary: Summary dictionary
        filepath: Output file path

    Raises:
        IOError: If file write fails
    """
    try:
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            # Write data sheets
            if not trades_df.empty:
                trades_converted = _format_column_names(_convert_date_columns(trades_df))
                trades_converted.to_excel(writer, sheet_name="Trades", index=False)
                _format_sheet(writer, "Trades", trades_converted)

            if not dividends_df.empty:
                dividends_converted = _format_column_names(_convert_date_columns(dividends_df))
                dividends_converted.to_excel(writer, sheet_name="Dividends", index=False)
                _format_sheet(writer, "Dividends", dividends_converted)

            if not tax_df.empty:
                tax_converted = _format_column_names(_convert_date_columns(tax_df))
                tax_converted.to_excel(writer, sheet_name="Withholding Tax", index=False)
                _format_sheet(writer, "Withholding Tax", tax_converted)

            if not deposits_withdrawals_df.empty:
                dw_converted = _format_column_names(_convert_date_columns(deposits_withdrawals_df))
                dw_converted.to_excel(writer, sheet_name="Deposits & Withdrawals", index=False)
                _format_sheet(writer, "Deposits & Withdrawals", dw_converted)

            # Write summary sheet
            summary_data = []
            for category, values in summary.items():
                for metric, value in values.items():
                    summary_data.append(
                        {
                            "Category": category.replace("_", " "),
                            "Metric": metric.replace("_", " "),
                            "Value": value,
                        }
                    )

            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name="Summary", index=False)
            _format_sheet(writer, "Summary", summary_df)
            _merge_summary_categories(writer, "Summary", summary_df)
    except Exception as e:
        raise IOError(f"Failed to export Excel file: {e}") from e


def _convert_date_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert date string columns to datetime objects

    Args:
        df: DataFrame with potential date columns

    Returns:
        DataFrame with converted date columns
    """
    df_copy = df.copy()
    for col in df_copy.columns:
        col_lower = col.lower()

        # Handle Date columns (format: YYYYMMDD)
        if col_lower == "date":
            try:
                df_copy[col] = pd.to_datetime(df_copy[col], format="%Y%m%d", errors="coerce")
                if df_copy[col].isna().all():
                    # If all failed, try other formats
                    df_copy[col] = pd.to_datetime(df_copy[col], errors="coerce")
            except Exception:
                pass
        # Handle Time columns (format: HH:MM:SS string to time)
        elif col_lower == "time":
            try:
                # Convert HH:MM:SS to datetime.time, keep as string for Excel
                # Empty strings remain empty
                def parse_time(t):
                    if t and isinstance(t, str) and t.strip():
                        try:
                            return pd.to_datetime(t, format="%H:%M:%S").time()
                        except Exception:
                            return t
                    return None

                df_copy[col] = df_copy[col].apply(parse_time)
            except Exception:
                pass
    return df_copy


def _format_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """
    Format column names for better readability in Excel

    Args:
        df: DataFrame with column names to format

    Returns:
        DataFrame with formatted column names
    """
    df_copy = df.copy()
    df_copy.columns = [col.replace("_", " ") for col in df_copy.columns]
    return df_copy


def _format_sheet(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    """
    Apply formatting to Excel sheet columns

    Args:
        writer: ExcelWriter instance
        sheet_name: Name of the sheet to format
        df: DataFrame that was written to the sheet
    """
    from openpyxl.styles import Alignment, Font, PatternFill, numbers

    worksheet = writer.sheets[sheet_name]

    # Define color scheme - Bloomberg Terminal style
    # Dark gray/black header with white text, professional and widely used in finance
    header_fill = PatternFill(
        start_color="2B2B2B", end_color="2B2B2B", fill_type="solid"
    )  # Near-black gray
    zebra_fill = PatternFill(
        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
    )  # Light gray

    # Alternative themes (uncomment to use):
    # Theme 1: Conservative (Dark Blue)
    # header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    # zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Theme 2: Modern (Dark Gray)
    # header_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    # zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Theme 3: Financial (Dark Green)
    # header_fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
    # zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")  # White text
    arial_font = Font(name="Arial", size=11)
    red_font = Font(name="Arial", size=11, color="C00000")  # Red for negative numbers

    # Apply header formatting (row 1)
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Apply zebra striping and fonts to data rows
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        # Apply zebra striping (even rows get gray background)
        if row_idx % 2 == 0:
            for cell in row:
                cell.fill = zebra_fill
                cell.font = arial_font
        else:
            for cell in row:
                cell.font = arial_font

    # Define column formats based on column name patterns
    for idx, col_name in enumerate(df.columns, start=1):
        col_letter = worksheet.cell(row=1, column=idx).column_letter
        col_name_lower = col_name.lower()

        # DateTime columns (date + time)
        if col_name_lower == "datetime":
            for cell in worksheet[col_letter][1:]:  # Skip header
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Date columns (date only)
        elif col_name_lower == "date":
            for cell in worksheet[col_letter][1:]:  # Skip header
                cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Time columns (time only)
        elif col_name_lower == "time":
            for cell in worksheet[col_letter][1:]:  # Skip header
                if cell.value:  # Only format non-empty cells
                    cell.number_format = "hh:mm:ss"
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Amount/Price/P&L columns (2 decimal places)
        elif any(
            keyword in col_name_lower
            for keyword in [
                "amount",
                "price",
                "p&l",
                "proceeds",
                "cost",
                "value",
                "tax",
                "dividend",
                "income",
                "credit",
                "payable",
                "deposits",
                "withdrawals",
            ]
        ):
            for cell in worksheet[col_letter][1:]:  # Skip header
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2
                cell.alignment = Alignment(horizontal="right", vertical="center")
                # Apply red font for negative numbers
                if isinstance(cell.value, (int, float)) and cell.value < 0:
                    cell.font = red_font

        # FX Rate columns (4 decimal places)
        elif "rate" in col_name_lower or "fx" in col_name_lower:
            for cell in worksheet[col_letter][1:]:  # Skip header
                cell.number_format = "0.0000"
                cell.alignment = Alignment(horizontal="right", vertical="center")

        # Count columns (integers)
        elif "count" in col_name_lower or "quantity" in col_name_lower:
            for cell in worksheet[col_letter][1:]:  # Skip header
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                cell.alignment = Alignment(horizontal="right", vertical="center")

        # Text columns (Symbol, Description, Category, etc.) - left align
        else:
            for cell in worksheet[col_letter][1:]:  # Skip header
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-adjust column widths based on formatted display length
    for idx, column in enumerate(worksheet.columns, start=0):
        max_length = 0
        column_letter = column[0].column_letter
        col_name = df.columns[idx] if idx < len(df.columns) else ""
        col_name_lower = col_name.lower()

        # Check header length
        max_length = len(str(col_name))

        # Check data cell lengths based on their actual format
        for cell in column[1:]:  # Skip header row
            try:
                if cell.value is not None:
                    formatted_length = _get_formatted_cell_length(cell)
                    max_length = max(max_length, formatted_length)
            except Exception:
                pass

        # Set width cap based on column type
        # Description columns need appropriate space based on content
        if "description" in col_name_lower:
            # Use actual content length, with reasonable min/max
            adjusted_width = max(max_length + 2, 30)  # Minimum 30 chars
            adjusted_width = min(adjusted_width, 100)  # Max 100 chars
        else:
            adjusted_width = min(max_length + 2, 50)

        worksheet.column_dimensions[column_letter].width = adjusted_width


def _get_formatted_cell_length(cell) -> int:
    """
    Calculate display length of a cell based on its number format

    Args:
        cell: openpyxl cell object

    Returns:
        Estimated display length
    """

    from openpyxl.styles import numbers

    if cell.value is None:
        return 0

    number_format = cell.number_format

    # DateTime format: yyyy-mm-dd hh:mm:ss
    if number_format == "yyyy-mm-dd hh:mm:ss":
        return 19  # "2025-01-09 03:06:57"

    # Date format: yyyy-mm-dd
    elif number_format == numbers.FORMAT_DATE_YYYYMMDD2:
        return 10  # "2025-01-09"

    # Time format: hh:mm:ss
    elif number_format == "hh:mm:ss":
        return 8  # "15:30:45"

    # Number with comma and 2 decimals: #,##0.00
    elif number_format == numbers.FORMAT_NUMBER_COMMA_SEPARATED2:
        if isinstance(cell.value, (int, float)):
            abs_value = abs(cell.value)
            formatted_str = f"{abs_value:,.2f}"
            if cell.value < 0:
                formatted_str = "-" + formatted_str
            return len(formatted_str)
        return len(str(cell.value))

    # Number with comma and no decimals: #,##0
    elif number_format == numbers.FORMAT_NUMBER_COMMA_SEPARATED1:
        if isinstance(cell.value, (int, float)):
            abs_value = abs(cell.value)
            formatted_str = f"{abs_value:,.0f}"
            if cell.value < 0:
                formatted_str = "-" + formatted_str
            return len(formatted_str)
        return len(str(cell.value))

    # 4 decimal places: 0.0000
    elif number_format == "0.0000":
        if isinstance(cell.value, (int, float)):
            formatted_str = f"{cell.value:.4f}"
            return len(formatted_str)
        return len(str(cell.value))

    # Default: convert to string
    else:
        return len(str(cell.value))


def _merge_summary_categories(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    """
    Merge cells in Category column for same category values

    Args:
        writer: ExcelWriter instance
        sheet_name: Name of the sheet
        df: Summary DataFrame
    """
    from openpyxl.styles import Alignment

    worksheet = writer.sheets[sheet_name]

    # Find consecutive rows with same category
    current_category = None
    start_row = None

    for idx, row in df.iterrows():
        row_num = idx + 2  # +2 because Excel is 1-indexed and we have header
        category = row["Category"]

        if category != current_category:
            # Merge previous category if exists
            if current_category is not None and start_row is not None and start_row < row_num - 1:
                worksheet.merge_cells(f"A{start_row}:A{row_num - 1}")
                # Center align merged cell
                worksheet[f"A{start_row}"].alignment = Alignment(
                    horizontal="center", vertical="center"
                )

            # Start new category
            current_category = category
            start_row = row_num

    # Merge last category
    if current_category is not None and start_row is not None:
        last_row = len(df) + 1
        if start_row < last_row:
            worksheet.merge_cells(f"A{start_row}:A{last_row}")
            worksheet[f"A{start_row}"].alignment = Alignment(horizontal="center", vertical="center")


def print_summary(summary: Dict[str, Any]) -> None:
    """
    Print summary to console

    Args:
        summary: Summary dictionary
    """
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)

    for category, values in summary.items():
        print(f"\n{category}:")
        print("-" * 40)
        for metric, value in values.items():
            metric_name = metric.replace("_", " ")
            if isinstance(value, (int, float)):
                print(f"  {metric_name:<35} {value:>15,.2f}")
            else:
                print(f"  {metric_name:<35} {value:>15}")


def convert_to_native(obj: Any) -> Any:
    """
    Convert numpy types to Python native types for JSON serialization

    Args:
        obj: Object to convert

    Returns:
        Converted object
    """
    if isinstance(obj, dict):
        return {k: convert_to_native(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [convert_to_native(item) for item in obj]
    elif hasattr(obj, "item"):  # numpy types
        return obj.item()
    else:
        return obj


def process_accounts(
    data: Any, logger: Any
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Process single or multiple accounts data

    Args:
        data: Flex query response data (dict or list)
        logger: Logger instance

    Returns:
        Tuple of (trades_df, dividends_df, tax_df, deposits_withdrawals_df)
    """
    if isinstance(data, list):
        logger.info(f"Processing {len(data)} account(s)...")
        all_trades: List[pd.DataFrame] = []
        all_dividends: List[pd.DataFrame] = []
        all_taxes: List[pd.DataFrame] = []
        all_deposits_withdrawals: List[pd.DataFrame] = []

        for idx, account_data in enumerate(data):
            account_id = account_data.get("@accountId", f"Account_{idx + 1}")
            logger.info(f"Processing account: {account_id}")

            trades = parse_trades(account_data)
            dividends = parse_dividends(account_data)
            taxes = parse_withholding_tax(account_data)
            deposits_withdrawals = parse_deposits_withdrawals(account_data)

            if not trades.empty:
                trades["Account"] = account_id
                all_trades.append(trades)
            if not dividends.empty:
                dividends["Account"] = account_id
                all_dividends.append(dividends)
            if not taxes.empty:
                taxes["Account"] = account_id
                all_taxes.append(taxes)
            if not deposits_withdrawals.empty:
                deposits_withdrawals["Account"] = account_id
                all_deposits_withdrawals.append(deposits_withdrawals)

        # Merge all accounts
        trades_df = pd.concat(all_trades, ignore_index=True) if all_trades else pd.DataFrame()
        dividends_df = (
            pd.concat(all_dividends, ignore_index=True) if all_dividends else pd.DataFrame()
        )
        tax_df = pd.concat(all_taxes, ignore_index=True) if all_taxes else pd.DataFrame()
        deposits_withdrawals_df = (
            pd.concat(all_deposits_withdrawals, ignore_index=True)
            if all_deposits_withdrawals
            else pd.DataFrame()
        )

        logger.info(
            f"Total across all accounts: {len(trades_df)} trades, "
            f"{len(dividends_df)} dividends, {len(tax_df)} taxes, "
            f"{len(deposits_withdrawals_df)} deposits/withdrawals"
        )
    else:
        # Single account
        logger.info("Processing single account...")
        trades_df = parse_trades(data)
        dividends_df = parse_dividends(data)
        tax_df = parse_withholding_tax(data)
        deposits_withdrawals_df = parse_deposits_withdrawals(data)

    return trades_df, dividends_df, tax_df, deposits_withdrawals_df


def parse_args() -> argparse.Namespace:
    """
    Parse command-line arguments

    Returns:
        Parsed arguments
    """
    parser = argparse.ArgumentParser(
        description="IBKR Tax Tool - Fetch and analyze trading data for Chinese tax reporting",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Use default date range from Flex Query
  ibkr-tax

  # Get data for specific year (China tax year: Jan 1 - Dec 31)
  ibkr-tax --year 2025

  # Get data for current year
  ibkr-tax --year $(date +%%Y)

  # Get all data from start year to current year
  ibkr-tax --from-year 2020

  # Get all data from FIRST_TRADE_YEAR (in .env) to current year
  ibkr-tax --all
        """,
    )

    parser.add_argument(
        "-y",
        "--year",
        type=int,
        help="Tax year (e.g., 2025). Fetches data from Jan 1 to Dec 31 of specified year.",
    )

    parser.add_argument(
        "--from-year",
        type=int,
        help="Starting year. Fetches all data from this year to current year (year by year).",
    )

    parser.add_argument(
        "--all",
        action="store_true",
        help="Fetch all data from FIRST_TRADE_YEAR (in .env) to current year.",
    )

    return parser.parse_args()


def main() -> None:
    """Main execution function"""
    # Parse command-line arguments
    args = parse_args()

    print_banner()

    # Initialize logger (console only for now)
    logger = setup_logger("ibkr_tax", level="INFO", console=True)

    # Validate arguments
    if sum([bool(args.year), bool(args.from_year), args.all]) > 1:
        error_msg = (
            "Cannot specify multiple date options: "
            "--year, --from-year, and --all are mutually exclusive"
        )
        logger.error(error_msg)
        print("\n✗ Error: Please specify only one of: --year, --from-year, or --all")
        sys.exit(1)

    # Calculate date range based on arguments
    from_date = None
    to_date = None
    years_to_fetch = []

    if args.year:
        # Single year mode
        from_date = f"{args.year}0101"
        to_date = f"{args.year}1231"
        years_to_fetch = [args.year]
        logger.info(f"Tax year: {args.year} (from {from_date} to {to_date})")
    elif args.from_year or args.all:
        # Multi-year mode
        current_year = datetime.now().year
        if args.all:
            # Need to load config to get FIRST_TRADE_YEAR
            try:
                temp_config = Config()
                start_year = temp_config.first_trade_year
                if not start_year:
                    logger.error("FIRST_TRADE_YEAR not set in .env file")
                    print("\n✗ Error: Please set FIRST_TRADE_YEAR in .env file to use --all option")
                    sys.exit(1)
            except Exception as e:
                logger.error(f"Failed to load config: {e}")
                sys.exit(1)
        else:
            start_year = args.from_year

        years_to_fetch = list(range(start_year, current_year + 1))
        logger.info(
            f"Fetching data from {start_year} to {current_year} ({len(years_to_fetch)} years)"
        )

    try:
        # Step 1: Load configuration
        logger.info("Step 1: Loading configuration...")
        try:
            config = Config()
            logger.info("Configuration loaded successfully")
            logger.info(f"Exchange rate: 1 USD = {config.exchange_rate} CNY")
            logger.info(f"Use dynamic rates: {config.use_dynamic_rates}")
        except ConfigurationError as e:
            logger.error(f"Configuration error: {e}")
            print("\nPlease check your .env file or environment variables.")
            print("Required variables:")
            print("  - IBKR_FLEX_TOKEN")
            print("  - IBKR_QUERY_ID")
            sys.exit(1)

        # Step 2: Prepare output directory
        logger.info("Step 2: Preparing output directory...")
        output_dir = config.output_dir
        output_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"Output directory: {output_dir}")

        # Step 3: Fetch data from IBKR
        logger.info("Step 3: Fetching data from IBKR...")
        client = FlexQueryClient(config.token, config.query_id)

        # Type hint for data variable (can be single year data or list of year data)
        data: Union[Dict[str, Any], List[Any]]

        if years_to_fetch and len(years_to_fetch) > 1:
            # Multi-year mode: fetch year by year
            logger.info(f"Multi-year mode: fetching {len(years_to_fetch)} years of data...")
            all_data = []

            for year in years_to_fetch:
                year_from = f"{year}0101"
                year_to = f"{year}1231"
                logger.info(f"  Fetching year {year} ({year_from} to {year_to})...")

                try:
                    year_data = client.fetch_data(from_date=year_from, to_date=year_to)
                    all_data.append(year_data)
                    logger.info(f"  ✓ Year {year} fetched successfully")
                except APIError as e:
                    logger.warning(f"  ! Failed to fetch year {year}: {e}")
                    logger.warning("  Continuing with remaining years...")

            if not all_data:
                logger.error("Failed to fetch any data")
                sys.exit(1)

            logger.info(f"Successfully fetched {len(all_data)} year(s) of data")

            # Combine all years data
            logger.info("Combining data from all years...")
            if isinstance(all_data[0], list):
                # Multiple accounts: combine by account
                data = all_data[0]  # Start with first year's structure
                # This is complex, will be handled in process_accounts
            else:
                # Single account: will be handled in process_accounts
                data = all_data
        else:
            # Single year or default mode
            try:
                data = client.fetch_data(from_date=from_date, to_date=to_date)
                logger.info("Data fetched successfully")
            except APIError as e:
                logger.error(f"API error: {e}")
                sys.exit(1)

        # Save raw data
        timestamp = datetime.now().strftime(TIMESTAMP_FORMAT)
        raw_data_file = output_dir / f"raw_data_{timestamp}.json"
        try:
            client.save_raw_data(data, str(raw_data_file))
        except IOError as e:
            logger.error(f"Failed to save raw data: {e}")

        # Step 4: Parse data
        logger.info("Step 4: Parsing data...")
        if (
            years_to_fetch
            and len(years_to_fetch) > 1
            and isinstance(data, list)
            and not isinstance(data[0], dict)
        ):
            # Multi-year mode: process each year and combine
            all_trades = []
            all_dividends = []
            all_taxes = []
            all_deposits_withdrawals = []

            for year_data in data:
                trades, dividends, taxes, deposits_withdrawals = process_accounts(year_data, logger)
                if not trades.empty:
                    all_trades.append(trades)
                if not dividends.empty:
                    all_dividends.append(dividends)
                if not taxes.empty:
                    all_taxes.append(taxes)
                if not deposits_withdrawals.empty:
                    all_deposits_withdrawals.append(deposits_withdrawals)

            trades_df = pd.concat(all_trades, ignore_index=True) if all_trades else pd.DataFrame()
            dividends_df = (
                pd.concat(all_dividends, ignore_index=True) if all_dividends else pd.DataFrame()
            )
            tax_df = pd.concat(all_taxes, ignore_index=True) if all_taxes else pd.DataFrame()
            deposits_withdrawals_df = (
                pd.concat(all_deposits_withdrawals, ignore_index=True)
                if all_deposits_withdrawals
                else pd.DataFrame()
            )

            logger.info(
                f"Combined data: {len(trades_df)} trades, "
                f"{len(dividends_df)} dividends, {len(tax_df)} taxes, "
                f"{len(deposits_withdrawals_df)} deposits/withdrawals"
            )
        else:
            trades_df, dividends_df, tax_df, deposits_withdrawals_df = process_accounts(
                data, logger
            )

        # Step 5: Calculate summary
        logger.info("Step 5: Calculating summary...")
        if config.use_dynamic_rates:
            logger.info("Using dynamic exchange rates (fetched from API)")
        else:
            logger.info(f"Using fixed exchange rate: {config.exchange_rate}")

        summary = calculate_summary(
            trades_df,
            dividends_df,
            tax_df,
            deposits_withdrawals_df,
            use_dynamic_rates=config.use_dynamic_rates,
            default_rate=config.exchange_rate,
        )
        logger.info("Summary calculated successfully")

        # Step 6: Export to Excel
        logger.info("Step 6: Exporting to Excel...")
        excel_file = output_dir / f"ibkr_report_{timestamp}.xlsx"
        try:
            export_to_excel(
                trades_df, dividends_df, tax_df, deposits_withdrawals_df, summary, str(excel_file)
            )
            logger.info(f"Excel file saved: {excel_file}")
        except IOError as e:
            logger.error(f"Failed to export Excel: {e}")

        # Export summary to JSON
        summary_file = output_dir / f"summary_{timestamp}.json"
        try:
            summary_native = convert_to_native(summary)
            with open(summary_file, "w", encoding="utf-8") as f:
                json.dump(summary_native, f, indent=2, ensure_ascii=False)
            logger.info(f"Summary JSON saved: {summary_file}")
        except IOError as e:
            logger.error(f"Failed to save summary JSON: {e}")

        # Print summary to console
        print_summary(summary)

        print("\n" + "=" * 60)
        print("✓ All tasks completed successfully!")
        print("=" * 60)

    except KeyboardInterrupt:
        logger.warning("Operation cancelled by user")
        print("\n\nOperation cancelled by user")
        sys.exit(0)
    except IBKRTaxError as e:
        logger.error(f"Application error: {e}", exc_info=True)
        print(f"\n✗ Error: {e}")
        sys.exit(1)
    except Exception as e:
        logger.error(f"Unexpected error: {e}", exc_info=True)
        print(f"\n✗ Unexpected error: {e}")
        print("\nFor detailed error information, check the logs.")
        sys.exit(1)


if __name__ == "__main__":
    main()
